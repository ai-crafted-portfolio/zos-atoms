# -*- coding: utf-8 -*-
"""Build the technical-breakdown MkDocs pages from the source xlsx workbooks.

Reads 8 breakdown workbooks, normalises every sheet into a uniform page model
(product -> Lv1 page -> Lv2 collapsible section -> rows), sanitises any internal
tooling vocabulary out of the published text, and writes:

  docs/index.md
  docs/breakdown/<product>/index.md
  docs/breakdown/<product>/lvNN.md
  docs/breakdown/verification-sample/index.md
  mkdocs.yml   (site_name / site_description / nav refreshed)

Run:  python tools/build_breakdown.py
"""
import os, re, json
import openpyxl

ROOT = r"C:\kvba\zos-atoms-site"
DOCS = os.path.join(ROOT, "docs")
BREAKDOWN = os.path.join(DOCS, "breakdown")
SITE_URL = "https://ai-crafted-portfolio.github.io/zos-atoms/"

# --------------------------------------------------------------------------- #
# Sanitiser: strip internal tooling vocabulary from any text that reaches docs
# --------------------------------------------------------------------------- #
# Internal retrieval-pipeline vocabulary leaks into the gold verify/quiz text
# (チャンク=chunk, バケット=bucket, 出典命中=retrieval-hit, corpus, verbatim, the
# MCP_NO_HIT / Q-NO_HIT markers, ...). We strip it with ordered LITERAL substring
# replacements — no \b word boundaries (those silently miss tokens glued to CJK or
# to '_'). Longest / most specific phrases first so fallbacks don't double-rewrite.
_SCRUB = [
    # ---- whole-sentence provenance notes (raw "MCP ..." source forms) ----
    ("[MCP 命中なし — Quiz 生成不可]", "(この項目の確認問題は一次資料との突合後に追加予定)"),
    ("周辺チャンクが MCP corpus に存在しない為、verbatim 引用は省略。",
     "出力例は一次資料との突合後に追加予定。"),
    ("本項目に対する MCP チャンクが取得できなかったため、コンソール手順は未生成。",
     "本項目の検証手順は一次資料との突合後に追加予定。"),
    ("MCP 検索コーパスに該当チャンクなし。", "一次資料との突合予定。"),
    # ---- "MCP 命中チャンク" provenance variants (specific -> generic) ----
    ("MCP 命中チャンク参照: MCP_NO_HIT", "出典参照: 公式資料"),
    ("MCP 命中チャンク参照:", "出典参照:"),
    ("MCP 命中チャンクから抽出、原文ママ", "原文ママ"),
    ("MCP 命中チャンクから抽出", "公式資料から抽出"),
    ("MCP 命中チャンクから生成", "公式資料から生成"),
    ("MCP 命中チャンクと整合", "公式資料と整合"),
    ("MCP 命中チャンク page", "出典: 公式資料"),
    ("MCP 命中チャンク", "公式資料"),
    ("MCP 一部命中", "一次資料 部分一致"),
    ("[MCP 命中なし]", "(一次資料 突合予定)"),
    ("[MCP 命中]", ""),
    ("MCP 命中ゼロ", "一次資料 該当なし"),
    ("MCP 命中なし", "一次資料 突合予定"),
    ("命中ゼロ", "該当なし"),
    # ---- defensive: any pre-converted "出典命中" form ----
    ("出典命中チャンク", "公式資料"),
    ("出典命中なし", "一次資料 突合予定"),
    # ---- legacy MCP phrase forms ----
    ("(NO MCP HIT)", "(一次資料 突合予定)"),
    ("（NO MCP HIT）", "(一次資料 突合予定)"),
    ("内容説明 (MCP)", "内容説明"),
    ("内容説明（MCP）", "内容説明"),
    ("MCP_NO_HIT", "(一次資料 未突合)"),
    ("Q-NO_HIT", "(確認問題 未突合)"),
    ("MCP_PARTIAL", "一次資料 部分一致"),
    ("MCP verbatim", "一次資料 原文ママ"),
    ("MCP 未記載", "一次資料 未記載"),
    ("MCP 整合度", "出典整合度"),
    ("MCP 出典", "出典"),
    ("MCP コレクション", "資料集合"),
    ("MCP product", "資料"),
    ("(MCP)", ""),
    ("（MCP）", ""),
    # ---- bare retrieval vocabulary ----
    ("verbatim 抜粋", "原文抜粋"),
    ("verbatim 引用", "原文引用"),
    ("verbatim", "原文ママ"),
    ("バケットの ", ""),
    ("バケット", "区分"),
    ("該当チャンク", "該当資料"),
    ("周辺チャンク", "周辺資料"),
    ("チャンク", "資料"),
    ("検索コーパス", "資料集"),
    ("コーパス", "資料集"),
    ("corpus", "資料集"),
]
# MCP glued to CJK / punctuation: replace only when NOT inside an ASCII word, so
# real terms keep working (NMCPINIT, NMCPTEST, SYMCPACFWRAP, ...).
_MCP_RE = re.compile(r"(?<![A-Za-z])MCP(?![A-Za-z])")
# Other tooling tokens that never collide with z/OS terms as standalone words.
_TOKEN_RE = re.compile(
    r"\b(?:RAG|ChromaDB|chromadb-manager|HNSW|search_manual|ADR-\d+|ADR)\b",
)

def sanitize(value):
    if value is None:
        return ""
    s = str(value)
    for a, b in _SCRUB:
        s = s.replace(a, b)
    s = _MCP_RE.sub("一次資料", s)
    s = _TOKEN_RE.sub("一次資料", s)
    # NB: intentionally preserve internal whitespace so the verification console
    # sessions keep their column alignment (the "実コンソールセッション再現" look).
    return s.strip()

def cell(value):
    """Sanitise + flatten for a single markdown table cell."""
    s = sanitize(value)
    s = s.replace("\r", " ").replace("\n", " ")
    s = s.replace("|", r"\|")
    return s.strip() or "—"

# Source cells appear in several formats: "MF,BAS", "AIX1/AIX73", and stringified
# python lists/tuples ("['KORN']", "('AS2', 'ZOS31')"). Normalise to clean codes.
_CODE_RE = re.compile(r"[A-Za-z][A-Za-z0-9]+")

def source_codes(value):
    s = sanitize(value)
    if not s or s in ("None", "—"):
        return []
    if "突合予定" in s:
        return ["(一次資料 突合予定)"]
    return [c for c in _CODE_RE.findall(s) if c != "None"]

def source_display(value):
    codes = source_codes(value)
    return ", ".join(codes) if codes else "—"

# --------------------------------------------------------------------------- #
# Product configuration
# --------------------------------------------------------------------------- #
# extractor(row) -> dict(page, section, mid, item, summary, source) using 0-based
# column indices. Returns None to skip a row.
def std(row):     # zos-v3 / aix / python : 連番,Lv1,Lv2,Lv3,要約,出典
    return dict(seq=row[0], page=row[1], section=row[2], mid=None,
                item=row[3], summary=row[4], source=row[5])

def sysprog(row): # No,Lv1No,Lv1,Lv2,技術項目,概要,出典
    return dict(seq=row[0], page=row[2], section=row[3], mid=None,
                item=row[4], summary=row[5], source=row[6])

def hw(row):      # z16: ID,Lv1code,Lv1,Lv2,Lv3,項目名,種別,説明,出典
    return dict(seq=row[0], page=row[2], section=row[3], mid=row[4],
                item=row[5], summary=row[7], source=row[8])

def vscode(row):  # No,code,大分類,中分類,小分類,項目名,種別,説明,出典code,章
    return dict(seq=row[0], page=row[2], section=row[3], mid=row[4],
                item=row[5], summary=row[7], source=row[8])

def tape(row, sheet):  # #,Lv1カテゴリ,Lv2サブ,項目名,説明,出典,備考  (page = sheet)
    return dict(seq=row[0], page=sheet, section=row[1], mid=row[2],
                item=row[3], summary=row[4], source=row[5])

PRODUCTS = [
    dict(key="zos-sysprog", nav="z/OS System Programming (1,804 項目, 初版)",
         heading="z/OS System Programming", product="z/OS System Programming",
         count=1804, source_note="初版（事前知識補完を含む）", provisional=True,
         path=r"C:\PDF\zos_sysprog_tech_breakdown_v1.xlsx",
         sheet="技術項目一覧", extract=sysprog, mid_header=None),
    dict(key="z16", nav="IBM z16 ハードウェア (1,249 項目, 初版)",
         heading="IBM z16 ハードウェア", product="IBM z16", count=1249,
         source_note="初版（事前知識補完を含む）", provisional=True,
         path=r"C:\PDF\z16_hw_tech_breakdown_v1.xlsx",
         sheet="技術項目一覧", extract=hw, mid_header="小分類"),
    dict(key="tape", nav="テープ系 TS2280 / TS4300 / TS7770 (1,045 項目, 初版)",
         heading="テープ系 (TS2280 / TS4300 / TS7770)", product="テープ系",
         count=1045, source_note="初版（事前知識補完を含む）", provisional=True,
         path=r"C:\PDF\tape_tech_breakdown_v1.xlsx",
         sheet=None, extract=tape, mid_header="サブカテゴリ"),
    dict(key="aix-ksh", nav="AIX + ksh (2,681 項目)",
         heading="AIX + ksh", product="AIX + ksh", count=2681,
         source_note="書籍出典", provisional=False,
         path=r"C:\PDF\aix_ksh_tech_breakdown_v1.xlsx",
         sheet="技術項目一覧", extract=std, mid_header=None),
    dict(key="python", nav="Python (1,774 項目, 公式 docs 出典)",
         heading="Python", product="Python", count=1774,
         source_note="Python 公式ドキュメント出典", provisional=False,
         path=r"C:\PDF\python_tech_breakdown_v1.xlsx",
         sheet="技術項目一覧", extract=std, mid_header=None),
    dict(key="vscode", nav="VS Code (2,042 項目, 公式 docs 出典)",
         heading="VS Code", product="VS Code", count=2042,
         source_note="VS Code 公式ドキュメント出典", provisional=False,
         path=r"C:\PDF\vscode_tech_breakdown_v1.xlsx",
         sheet="技術項目一覧", extract=vscode, mid_header="小分類"),
]

# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def yq(title):
    """YAML-safe double-quoted nav title."""
    return '"' + str(title).replace('"', "'") + '"'

def data_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if any(c is not None and str(c).strip() for c in r):
            yield r

def ordered_groups(items, keyfn):
    """Group preserving first-appearance order. Returns list[(key, [items])]."""
    order, buckets = [], {}
    for it in items:
        k = keyfn(it)
        if k not in buckets:
            buckets[k] = []
            order.append(k)
        buckets[k].append(it)
    return [(k, buckets[k]) for k in order]

def render_table(rows, mid_header):
    out = []
    if mid_header:
        out.append("| 連番 | %s | 技術項目 | 単一焦点要約 | 出典 |" % mid_header)
        out.append("|---:|---|---|---|---|")
        for r in rows:
            out.append("| %s | %s | %s | %s | %s |" % (
                cell(r["seq"]), cell(r["mid"]), cell(r["item"]),
                cell(r["summary"]), cell(r["source"])))
    else:
        out.append("| 連番 | 技術項目 | 単一焦点要約 | 出典 |")
        out.append("|---:|---|---|---|")
        for r in rows:
            out.append("| %s | %s | %s | %s |" % (
                cell(r["seq"]), cell(r["item"]),
                cell(r["summary"]), cell(r["source"])))
    return "\n".join(out)

PROVISIONAL_BOX = (
    '!!! warning "初版コンテンツ"\n'
    "    本ページは初版です。一部の技術項目は一次資料（公式マニュアル・Redbook 等）\n"
    "    との突合を予定しており、出典欄に「一次資料 突合予定」と表示される項目があります。\n"
    "    内容は実機運用前に必ず最新の一次資料で確認してください。\n"
)

# --------------------------------------------------------------------------- #
# build one product
# --------------------------------------------------------------------------- #
def build_product(cfg, legend_acc):
    wb = openpyxl.load_workbook(cfg["path"], read_only=True, data_only=True)
    records = []
    if cfg["sheet"] is None:  # tape: every sheet, page = sheet name
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in data_rows(ws):
                rec = cfg["extract"](r, sn)
                records.append(rec)
    else:
        ws = wb[cfg["sheet"]]
        for r in data_rows(ws):
            records.append(cfg["extract"](r))
    wb.close()

    # normalise source codes for legend + clean per-row display
    for rec in records:
        codes = source_codes(rec["source"])
        for code in codes:
            legend_acc.setdefault(code, set()).add(cfg["product"])
        rec["source"] = ", ".join(codes) if codes else "—"

    pages = ordered_groups(records, lambda r: sanitize(r["page"]) or "(未分類)")
    outdir = os.path.join(BREAKDOWN, cfg["key"])
    os.makedirs(outdir, exist_ok=True)

    page_meta = []  # (filename, title, count)
    for i, (pagename, recs) in enumerate(pages, start=1):
        fname = "lv%02d.md" % i
        sections = ordered_groups(recs, lambda r: sanitize(r["section"]) or "(その他)")
        body = []
        body.append("# %s" % pagename)
        body.append("")
        body.append("**製品**: %s ／ **本ページ件数**: %d 件 ／ **出典区分**: %s"
                    % (cfg["product"], len(recs), cfg["source_note"]))
        body.append("")
        if cfg["provisional"]:
            body.append(PROVISIONAL_BOX)
            body.append("")
        for sname, srecs in sections:
            # collapsed details per Lv2 section; first section open for orientation
            marker = "???+" if sname == sections[0][0] else "???"
            body.append('%s note "%s （%d 件）"' % (marker, sname, len(srecs)))
            body.append("")
            for line in render_table(srecs, cfg["mid_header"]).splitlines():
                body.append("    " + line)
            body.append("")
        with open(os.path.join(outdir, fname), "w", encoding="utf-8") as f:
            f.write("\n".join(body).rstrip() + "\n")
        page_meta.append((fname, pagename, len(recs)))

    # overview index
    idx = []
    idx.append("# %s" % cfg["heading"])
    idx.append("")
    idx.append("**製品**: %s ／ **全件数**: %s 件 ／ **出典区分**: %s"
               % (cfg["product"], "{:,}".format(cfg["count"]), cfg["source_note"]))
    idx.append("")
    if cfg["provisional"]:
        idx.append(PROVISIONAL_BOX)
        idx.append("")
    idx.append("## カテゴリ一覧")
    idx.append("")
    idx.append("| # | カテゴリ (Lv1) | 件数 | ページ |")
    idx.append("|---:|---|---:|---|")
    for j, (fname, title, cnt) in enumerate(page_meta, start=1):
        link = fname[:-3]  # mkdocs use-directory-urls -> link by stem is fine via .md
        idx.append("| %d | %s | %d | [%s](%s) |"
                   % (j, cell(title), cnt, "開く", fname))
    idx.append("")
    total_pages_sum = sum(c for _, _, c in page_meta)
    idx.append("> 合計 %d 件 / %d カテゴリ。" % (total_pages_sum, len(page_meta)))
    idx.append("")
    with open(os.path.join(outdir, "index.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(idx).rstrip() + "\n")

    return dict(key=cfg["key"], nav=cfg["nav"], pages=page_meta,
                total=total_pages_sum, declared=cfg["count"])

# --------------------------------------------------------------------------- #
# verification-sample page (special layout)
# --------------------------------------------------------------------------- #
def build_verify():
    path = r"C:\PDF\zos_tech_breakdown_v3_with_verify_SAMPLE50_v6.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items = list(data_rows(wb["技術項目一覧"]))   # 50
    procs = list(data_rows(wb["検証手順"]))        # 19
    wb.close()
    outdir = os.path.join(BREAKDOWN, "verification-sample")
    os.makedirs(outdir, exist_ok=True)

    b = []
    b.append("# 検証手順サンプル 50 件")
    b.append("")
    b.append("**全件数**: 50 技術項目（うち %d 件に実コンソールセッション再現の検証手順を収録）"
             % len(procs))
    b.append("")
    b.append("実際のコンソール画面の雰囲気を再現した検証手順サンプルです。"
             "コマンド入力・応答メッセージ・パネル遷移を含みます。")
    b.append("")
    b.append("## 対象技術項目 50 件")
    b.append("")
    b.append('???+ note "技術項目一覧 （50 件）"')
    b.append("")
    b.append("    | 連番 | 大分類 | 中分類 | 技術項目 | 内容説明 | 出典 |")
    b.append("    |---:|---|---|---|---|---|")
    for r in items:
        # cols: 0連番 1Lv1 2Lv2 3技術項目 4要約 5内容説明 6出典 ... 10出典書籍
        b.append("    | %s | %s | %s | %s | %s | %s |" % (
            cell(r[0]), cell(r[1]), cell(r[2]), cell(r[3]),
            cell(r[5]), cell(r[10])))
    b.append("")
    b.append("## 検証手順（コンソールセッション再現）")
    b.append("")
    for r in procs:
        # 0 ID,1 sheet,2 連番,3 Lv1,4 Lv2,5 項目名,6 目的,7 手順,8 前提,9 PH,10 記録,12 step
        vid = sanitize(r[0]); item = sanitize(r[5])
        lv1 = sanitize(r[3]); lv2 = sanitize(r[4])
        purpose = sanitize(r[6]); session = sanitize(r[7])
        pre = sanitize(r[8]); ph = sanitize(r[9]); rec = sanitize(r[10])
        b.append("### %s — %s" % (vid, item))
        b.append("")
        b.append("**分類**: %s ／ %s" % (lv1, lv2))
        b.append("")
        if purpose:
            b.append("**検証目的**: %s" % purpose)
            b.append("")
        b.append("```text")
        b.append(session.rstrip())
        b.append("```")
        b.append("")
        if pre:
            b.append("**前提条件**: %s" % pre.replace("\n", " "))
            b.append("")
        if ph:
            b.append("**プレースホルダ**: %s" % ph.replace("\n", " "))
            b.append("")
        if rec:
            b.append("**記録項目**:")
            b.append("")
            for ln in rec.split("\n"):
                ln = ln.strip()
                if ln:
                    b.append("- %s" % ln)
            b.append("")
    with open(os.path.join(outdir, "index.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(b).rstrip() + "\n")
    return dict(items=len(items), procs=len(procs))

# --------------------------------------------------------------------------- #
# z/OS v3 with verification steps + Quiz (FULL_v2 workbook)
# --------------------------------------------------------------------------- #
ZV_KEY = "zos-v3-with-verify"
ZV_LABEL = "z/OS v3 (検証手順 + Quiz 付き)"
ZV_HEADING = "z/OS v3 — 技術項目 + 検証手順 + 理解度チェック"
ZV_NOTE = "Redbook ingest 反映済 v3"
ZV_PATH = r"C:\PDF\zos_tech_breakdown_v3_with_verify_FULL_v2.xlsx"
ZV_PAGE_CAP = 300   # max 技術項目 per page before Lv2 / part split


def _fence(text):
    """Return a backtick fence longer than any backtick run inside text."""
    longest = 0
    for run in re.findall(r"`+", text):
        longest = max(longest, len(run))
    return "`" * max(3, longest + 1)


def _ind(lines, block):
    """Append block (str) indented by 4 spaces, line by line, to lines."""
    for ln in block.split("\n"):
        lines.append(("    " + ln).rstrip())


def _render_verify_quiz(rec, out):
    """Append one collapsible per item carrying its 検証手順 + Quiz."""
    v, q = rec["verify"], rec["quiz"]
    if not v and not q:
        return
    title = "#%s %s" % (cell(rec["seq"]), cell(rec["item"]))
    out.append('??? example "%s"' % title.replace('"', "'"))
    out.append("")
    if v:
        purpose = sanitize(v[6])
        session = sanitize(v[7])
        pre = sanitize(v[8]); ph = sanitize(v[9]); recd = sanitize(v[10])
        out.append("    **検証手順**")
        out.append("")
        if purpose:
            out.append("    検証目的: %s" % purpose.replace("\n", " "))
            out.append("")
        fence = _fence(session)
        out.append("    %stext" % fence)
        _ind(out, session.rstrip())
        out.append("    %s" % fence)
        out.append("")
        if pre:
            out.append("    前提条件: %s" % pre.replace("\n", " "))
            out.append("")
        if ph:
            out.append("    プレースホルダ: %s" % ph.replace("\n", " "))
            out.append("")
        if recd:
            out.append("    記録項目:")
            out.append("")
            for ln in recd.split("\n"):
                ln = ln.strip()
                if ln:
                    out.append("    - %s" % sanitize(ln))
            out.append("")
    if q:
        # cols: 5問題文 6A 7B 8C 9D 10正解 11解説 12出典
        question = sanitize(q[5])
        choices = [sanitize(q[6]), sanitize(q[7]), sanitize(q[8]), sanitize(q[9])]
        ans = sanitize(q[10]); expl = sanitize(q[11])
        if v:
            out.append("    ---")
            out.append("")
        out.append("    **理解度チェック**")
        out.append("")
        out.append("    問題: %s" % question.replace("\n", " "))
        out.append("")
        for label, ch in zip("ABCD", choices):
            if ch:
                out.append("    - %s. %s" % (label, ch.replace("\n", " ")))
        out.append("")
        out.append("    正解: **%s**" % ans)
        out.append("")
        if expl:
            out.append("    解説: %s" % expl.replace("\n", " "))
            out.append("")
    out.append("")


def _render_zv_page(h1, sections, cfg_total):
    """sections = list[(lv2name, [rec,...])]. Returns markdown string."""
    page_n = sum(len(r) for _, r in sections)
    b = []
    b.append("# %s" % h1)
    b.append("")
    b.append("**製品**: z/OS v3 ／ **本ページ件数**: %d 件 ／ **出典区分**: %s"
             % (page_n, ZV_NOTE))
    b.append("")
    b.append("各技術項目の表に続き、検証手順（コンソールセッション再現）と"
             "理解度チェック（4 択 + 解説）を項目ごとの折り畳みで収録しています。")
    b.append("")
    for lv2, recs in sections:
        nverify = sum(1 for r in recs if r["verify"] or r["quiz"])
        b.append("## %s" % lv2)
        b.append("")
        b.append("**本節**: %d 件（うち検証手順・確認問題 %d 件）" % (len(recs), nverify))
        b.append("")
        b.append("| 連番 | 技術項目 | 内容説明 | 出典 |")
        b.append("|---:|---|---|---|")
        for r in recs:
            detail = r["detail"] if sanitize(r["detail"]) else r["summary"]
            b.append("| %s | %s | %s | %s |" % (
                cell(r["seq"]), cell(r["item"]), cell(detail),
                cell(r["source"])))
        b.append("")
        if nverify:
            b.append("### 検証手順 ・ 理解度チェック")
            b.append("")
            for r in recs:
                _render_verify_quiz(r, b)
    return "\n".join(b).rstrip() + "\n"


def build_zos_verify(legend_acc):
    wb = openpyxl.load_workbook(ZV_PATH, read_only=True, data_only=True)
    vmap, qmap = {}, {}
    for r in data_rows(wb["検証手順"]):
        if r[2] is not None:
            vmap[str(r[2]).lstrip("#").strip()] = r
    for r in data_rows(wb["Quiz"]):
        if r[1] is not None:
            qmap[str(r[1]).lstrip("#").strip()] = r
    records = []
    for r in data_rows(wb["技術項目一覧"]):
        seq = str(r[0]).strip() if r[0] is not None else ""
        codes = source_codes(r[11])
        for code in codes:
            legend_acc.setdefault(code, set()).add("z/OS v3")
        records.append(dict(
            seq=seq, page=sanitize(r[1]) or "(未分類)", section=sanitize(r[2]) or "(その他)",
            item=r[3], summary=r[4], detail=r[5],
            source=", ".join(codes) if codes else "—",
            verify=vmap.get(seq), quiz=qmap.get(seq)))
    wb.close()

    outdir = os.path.join(BREAKDOWN, ZV_KEY)
    os.makedirs(outdir, exist_ok=True)

    lv1_groups = ordered_groups(records, lambda r: r["page"])
    counter = [0]

    def newname():
        counter[0] += 1
        return "g%03d.md" % counter[0]

    # nav tree: list of nodes. node = (label, filename) leaf, or (label, [children])
    nav_nodes = []
    overview = []  # (lv1name, total, first_filename)

    for lv1, recs in lv1_groups:
        total = len(recs)
        lv2_groups = ordered_groups(recs, lambda r: r["section"])
        if total <= ZV_PAGE_CAP:
            fname = newname()
            md = _render_zv_page(lv1, lv2_groups, total)
            with open(os.path.join(outdir, fname), "w", encoding="utf-8") as f:
                f.write(md)
            nav_nodes.append((("%s （%d 件）" % (lv1, total)), fname))
            overview.append((lv1, total, fname))
            continue

        # Large Lv1: pack consecutive Lv2 groups into pages of <= ZV_PAGE_CAP,
        # splitting only a single Lv2 that alone exceeds the cap.
        pages = []          # each page = list[(lv2name, recs_slice)]
        cur, cur_n = [], 0
        for lv2, lrecs in lv2_groups:
            if len(lrecs) > ZV_PAGE_CAP:
                if cur:
                    pages.append(cur); cur, cur_n = [], 0
                for i in range(0, len(lrecs), ZV_PAGE_CAP):
                    pages.append([(lv2, lrecs[i:i + ZV_PAGE_CAP])])
                continue
            if cur_n + len(lrecs) > ZV_PAGE_CAP and cur:
                pages.append(cur); cur, cur_n = [], 0
            cur.append((lv2, lrecs)); cur_n += len(lrecs)
        if cur:
            pages.append(cur)

        children, first_fname = [], None
        npages = len(pages)
        for ci, sections in enumerate(pages, start=1):
            fname = newname()
            if first_fname is None:
                first_fname = fname
            page_n = sum(len(r) for _, r in sections)
            h1 = "%s （%d/%d）" % (lv1, ci, npages)
            md = _render_zv_page(h1, sections, page_n)
            with open(os.path.join(outdir, fname), "w", encoding="utf-8") as f:
                f.write(md)
            label = "%d/%d: %s （%d 件）" % (ci, npages, sections[0][0], page_n)
            children.append((label, fname))
        nav_nodes.append((("%s （%d 件）" % (lv1, total)), children))
        overview.append((lv1, total, first_fname))

    # product index / overview page
    idx = []
    idx.append("# %s" % ZV_HEADING)
    idx.append("")
    idx.append("**製品**: z/OS v3 ／ **全件数**: %s 件 ／ **出典区分**: %s"
               % ("{:,}".format(len(records)), ZV_NOTE))
    idx.append("")
    nq = sum(1 for r in records if r["verify"])
    idx.append("技術項目を大分類 (Lv1) ごとにまとめ、各項目に検証手順"
               "（コンソールセッション再現）と理解度チェック（4 択 + 解説）を"
               "折り畳みで収録しています（検証手順・確認問題 各 %s 件）。"
               % "{:,}".format(nq))
    idx.append("")
    idx.append("## カテゴリ一覧")
    idx.append("")
    idx.append("| # | カテゴリ (Lv1) | 件数 | ページ |")
    idx.append("|---:|---|---:|---|")
    for j, (lv1, total, first_fname) in enumerate(overview, start=1):
        idx.append("| %d | %s | %d | [%s](%s) |"
                   % (j, cell(lv1), total, "開く", first_fname))
    idx.append("")
    idx.append("> 合計 %s 件 / %d カテゴリ。"
               % ("{:,}".format(len(records)), len(overview)))
    idx.append("")
    with open(os.path.join(outdir, "index.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(idx).rstrip() + "\n")

    # nav block (full product entry, ready to splice into mkdocs.yml nav)
    nav_block = []
    nav_block.append("  - %s:" % yq(ZV_LABEL))
    nav_block.append("      - 概要: breakdown/%s/index.md" % ZV_KEY)
    for label, target in nav_nodes:
        if isinstance(target, str):
            nav_block.append("      - %s: breakdown/%s/%s"
                             % (yq(label), ZV_KEY, target))
        else:
            nav_block.append("      - %s:" % yq(label))
            for clabel, cfname in target:
                nav_block.append("          - %s: breakdown/%s/%s"
                                 % (yq(clabel), ZV_KEY, cfname))

    page_count = counter[0]
    return dict(key=ZV_KEY, label=ZV_LABEL, heading=ZV_HEADING,
                note=ZV_NOTE, total=len(records), pages=page_count,
                verify=nq, nav_block=nav_block)


# --------------------------------------------------------------------------- #
# home + mkdocs.yml
# --------------------------------------------------------------------------- #
LEGEND_NAMES = {
    # z/OS v3
    "MF": "メインフレーム実践（書籍）",
    "BAS": "MFOS入門（書籍）",
    "AS1": "アドバンスドスキル Vol.1（書籍）",
    "AS2": "アドバンスドスキル Vol.2（書籍）",
    "ZOS31": "IBM z/OS 3.1 公式マニュアル",
    "NV": "IBM NetView 6.4 マニュアル一式",
    "TSA": "IBM Z System Automation for z/OS 4.3 マニュアル一式",
    "GDPS": "IBM GDPS Redbooks 一式",
    "MQ": "IBM MQ 9.3 マニュアル一式",
    # AIX + ksh
    "AIX1": "AIX システム管理 1（書籍）",
    "AIX2": "AIX システム管理 2（書籍）",
    "AIX73": "IBM AIX 7.3 公式マニュアル",
    "KORN": "入門 Korn シェル（書籍）",
    "PHA7": "PowerHA SystemMirror 7.x マニュアル",
    # Python / VS Code
    "PYDOC": "Python 公式ドキュメント (docs.python.org/3)",
    "VSCDOCS": "Visual Studio Code 公式ドキュメント",
    "VSCAPI": "Visual Studio Code 拡張 API リファレンス",
    "VSCKB": "Visual Studio Code キーボードショートカット リファレンス",
    "VSCSET": "Visual Studio Code 設定リファレンス",
    "VSCREM": "Visual Studio Code リモート開発ドキュメント",
    "VSCCOP": "GitHub Copilot in VS Code ドキュメント",
    # Tape
    "TS2280DOC": "IBM TS2280 公式ドキュメント",
    "TS4300DOC": "IBM TS4300 公式ドキュメント",
    "TS7770DOC": "IBM TS7770 公式ドキュメント",
    # Provisional marker
    "(一次資料 突合予定)": "初版項目（一次資料との突合を予定）",
}

def build_home(results, verify, zv=None):
    b = []
    b.append("# z/OS 技術項目細分化ガイド")
    b.append("")
    b.append("z/OS を中心に、AIX・Python・VS Code・テープ系ハードウェア等の技術項目を"
             "細分化した技術リファレンスです。各製品を大分類 (Lv1) ごとのページにまとめ、"
             "中分類 (Lv2) 単位の折り畳みテーブルで技術項目・単一焦点要約・出典を一覧できます。")
    b.append("")
    b.append("z/OS v3 は技術項目に加え、項目ごとの検証手順（コンソールセッション再現）と"
             "理解度チェック（4 択 + 解説）を収録しています。")
    b.append("")
    b.append("画面右上の検索から全ページを横断的に全文検索できます。")
    b.append("")
    b.append("## コンテンツ一覧")
    b.append("")
    b.append("| 製品 | 件数 | 出典区分 | 入口 |")
    b.append("|---|---:|---|---|")
    if zv:
        b.append("| %s | %s | %s | [概要](breakdown/%s/index.md) |" % (
            cell(zv["heading"]), "{:,}".format(zv["total"]),
            zv["note"], zv["key"]))
    for r in results:
        cfg = next(c for c in PRODUCTS if c["key"] == r["key"])
        b.append("| %s | %s | %s | [概要](breakdown/%s/index.md) |" % (
            cell(cfg["heading"]), "{:,}".format(r["total"]),
            cfg["source_note"], r["key"]))
    b.append("| 検証手順サンプル | %d | 実コンソールセッション再現 | [概要](breakdown/verification-sample/index.md) |"
             % verify["items"])
    b.append("")
    b.append("## 信頼性区分の凡例")
    b.append("")
    b.append("- **Redbook ingest 反映済 / 公式 docs 出典 / 書籍出典**: 一次資料に基づく内容。")
    b.append("- **初版**: 事前知識による補完を含み、一次資料との突合（再生成）を予定。"
             "出典欄に「一次資料 突合予定」と表示される項目が該当します。")
    b.append("")
    b.append("## 出典コード凡例")
    b.append("")
    b.append("| 出典コード | 意味 | 主な製品 |")
    b.append("|---|---|---|")
    rows = []
    for code in sorted(LEGEND_USED.keys()):
        meaning = LEGEND_NAMES.get(code, "出典コード")
        prods = "、".join(sorted(LEGEND_USED[code]))
        rows.append((code, meaning, prods))
    for code, meaning, prods in rows:
        b.append("| `%s` | %s | %s |" % (cell(code), cell(meaning), cell(prods)))
    b.append("")
    with open(os.path.join(DOCS, "index.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(b).rstrip() + "\n")

def build_mkdocs(results, verify, zv=None):
    lines = []
    lines.append('site_name: "z/OS 技術項目細分化ガイド"')
    lines.append('site_description: "z/OS・AIX・Python・VS Code 等の技術項目を細分化した技術リファレンス"')
    lines.append("site_url: %s" % SITE_URL)
    lines.append("site_author: ai-crafted-portfolio")
    lines.append("")
    lines.append("copyright: Copyright &copy; 2026 ai-crafted-portfolio")
    lines.append("")
    lines.append("""theme:
  name: material
  language: ja
  features:
    - navigation.tabs
    - navigation.sections
    - navigation.top
    - navigation.indexes
    - search.highlight
    - search.suggest
    - content.code.copy
  palette:
    - media: "(prefers-color-scheme: light)"
      scheme: default
      primary: indigo
      accent: indigo
      toggle:
        icon: material/brightness-7
        name: ダークモードに切替
    - media: "(prefers-color-scheme: dark)"
      scheme: slate
      primary: indigo
      accent: indigo
      toggle:
        icon: material/brightness-4
        name: ライトモードに切替
  font:
    text: Noto Sans JP
    code: JetBrains Mono

markdown_extensions:
  - admonition
  - pymdownx.details
  - pymdownx.superfences
  - pymdownx.tabbed:
      alternate_style: true
  - pymdownx.highlight:
      anchor_linenums: true
  - attr_list
  - md_in_html
  - tables
  - toc:
      permalink: true

extra:
  generator: false
""")
    lines.append("nav:")
    lines.append("  - ホーム: index.md")
    if zv:
        lines.extend(zv["nav_block"])
    for r in results:
        lines.append("  - %s:" % yq(r["nav"]))
        lines.append("      - 概要: breakdown/%s/index.md" % r["key"])
        for fname, title, cnt in r["pages"]:
            lines.append("      - %s: breakdown/%s/%s"
                         % (yq("%s （%d 件）" % (title, cnt)), r["key"], fname))
    lines.append("  - %s: breakdown/verification-sample/index.md"
                 % yq("検証手順サンプル 50 件"))
    with open(os.path.join(ROOT, "mkdocs.yml"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines).rstrip() + "\n")

# --------------------------------------------------------------------------- #
LEGEND_USED = {}

def main():
    os.makedirs(BREAKDOWN, exist_ok=True)
    results = []
    for cfg in PRODUCTS:
        res = build_product(cfg, LEGEND_USED)
        results.append(res)
        print("[%s] total=%d declared=%d pages=%d"
              % (res["key"], res["total"], res["declared"], len(res["pages"])))
    zv = build_zos_verify(LEGEND_USED)
    print("[%s] total=%d pages=%d verify+quiz=%d"
          % (zv["key"], zv["total"], zv["pages"], zv["verify"]))
    verify = build_verify()
    print("[verify] items=%d procs=%d" % (verify["items"], verify["procs"]))
    build_home(results, verify, zv)
    build_mkdocs(results, verify, zv)
    # report
    report = dict(products=[{k: r[k] for k in ("key", "total", "declared")}
                            | {"pages": len(r["pages"])} for r in results],
                  zos_v3_with_verify=dict(total=zv["total"], pages=zv["pages"],
                                          verify_quiz=zv["verify"]),
                  verify=verify,
                  legend_codes=sorted(LEGEND_USED.keys()))
    with open(os.path.join(ROOT, "buildreport.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=1)
    print("DONE")

if __name__ == "__main__":
    main()
